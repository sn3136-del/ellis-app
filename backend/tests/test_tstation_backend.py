"""The quality-control backend, pinned to Trip.com's acceptance standard:
25-field records with per-field status, combined filtering, the change log,
and the two-sheet Excel export."""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.visa_snapshot import kimi_primary, tstation

READER = {"authorization": "Bearer dev-token", "x-org-id": "org-a",
          "x-user-id": "reader-1"}
ADMIN = {"authorization": "Bearer admin-token", "x-org-id": "org-b",
         "x-user-id": "operator-1"}

ANSWER = {
    "disposition": "VISA_REQUIRED", "visa_category": "Tourist visa",
    "permitted_stay": "30 days", "passport_validity": "6 months",
    "required_documents": ["passport", "photo"],
    "application_channel": "EMBASSY_OR_CONSULATE",
    "government_fee": {"amount": 100, "currency": "USD"},
    "processing_time": "5 working days", "confidence": "high",
    "source_url": "https://www.mofa.go.jp/j_info/visit/visa/index.html",
    "visa_products": [
        {"type": "Single-entry tourist", "entry": "single",
         "validity": "3 months", "max_stay_days": 30,
         "fee": {"amount": 100, "currency": "USD"}, "notes": None},
        {"type": "Multiple-entry tourist", "entry": "multiple",
         "validity": "5 years", "max_stay_days": 90,
         "fee": {"amount": 250, "currency": "USD"}, "notes": "High income"},
    ],
}


@pytest.fixture()
def client():
    c = TestClient(app)
    yield c
    kimi_primary.set_provider(None)


def test_the_25_field_record_speaks_their_dictionary_exactly():
    route = {"passport_nationality": "CHN", "destination_country": "JPN",
             "travel_purpose": "tourism"}
    rows = tstation.records_for_route(route, ANSWER, None, "2026-08-27T00:00:00")
    assert len(rows) == 2                      # one record per visa product
    r = rows[0]
    assert set(tstation.FIELD_ORDER) <= set(r.keys())
    assert r["visa_type_name"] == "Single-entry tourist"
    assert (r["validity_duration"], r["validity_unit"]) == (3, "Month")
    assert (r["max_stay_duration"], r["max_stay_unit"]) == (30, "Day")
    assert r["entries"] == "Single"
    assert (r["visa_fee_amount"], r["visa_fee_currency"]) == (100, "USD")
    assert r["application_method"] == "Embassy Submission"
    assert (r["processing_min_days"], r["processing_unit"]) == (5, "Working Day")
    # The spec ladder. An answer that asserts visa PRODUCTS but was never
    # checked against its official page is Low however good its URL looks:
    # an audit of every such record found 19 of 21 wrong (superseded fees,
    # products the destination does not issue, visas demanded of exempt
    # travellers). Once the official page has been read and agrees, the same
    # answer is Medium.
    assert r["confidence_level"] == "Low"
    ok = tstation.records_for_route(route, ANSWER, None, "2026-08-27T00:00:00",
                                    grounded_ok=True)
    assert ok[0]["confidence_level"] == "Medium"
    bare = {k: v for k, v in ANSWER.items()
            if k not in ("source_url", "official_portal_url")}
    low = tstation.records_for_route(route, bare, None, "2026-08-27T00:00:00")
    assert low[0]["confidence_level"] == "Low"
    r2 = rows[1]
    assert (r2["validity_duration"], r2["validity_unit"]) == (5, "Year")
    assert r2["entries"] == "Multiple"
    assert r2["special_conditions"] == "High income"


def test_visa_free_yields_one_clean_record_and_human_check_is_high():
    route = {"passport_nationality": "SGP", "destination_country": "CHN",
             "travel_purpose": "tourism"}
    g = {"disposition": "VISA_EXEMPT", "permitted_stay": "30 days",
         "permitted_stay_days": 30, "confidence": "high"}
    prov = {"source_url": "https://cs.mfa.gov.cn/x", "verified_at": "2026-08-22",
            "verified_by": "Ellis source audit"}
    rows = tstation.records_for_route(route, g, prov,
                                      valid_until="2026-11-20T00:00:00")
    assert len(rows) == 1
    r = rows[0]
    assert r["visa_requirement"] == "Visa-free"
    assert r["visa_type_name"] == "No visa needed"
    assert r["visa_fee_amount"] == 0
    assert r["confidence_level"] == "High"     # a person verified it
    assert r["source_url"] == "https://cs.mfa.gov.cn/x"
    assert r["collected_at"] == "2026-08-22"
    assert tstation.completeness(r) == 1.0


def test_field_status_reports_missing_required_fields():
    route = {"passport_nationality": "CHN", "destination_country": "JPN",
             "travel_purpose": "tourism"}
    rows = tstation.records_for_route(
        route, {"disposition": "VISA_REQUIRED", "confidence": "high"}, None)
    st = tstation.field_status(rows[0])
    assert st["visa_fee_amount"] == "missing"
    assert st["consulate_district"] == "optional-empty"
    assert st["visa_requirement"] == "filled"
    assert tstation.completeness(rows[0]) < 1.0


def _warm(client, nat, dest):
    kimi_primary.set_provider(lambda system, user: dict(ANSWER))
    r = client.post("/database/lookup", headers=READER,
                    json={"nationality": nat, "destination": dest})
    assert r.status_code == 200


def test_records_endpoint_filters_and_checklists(client):
    _warm(client, "NZL", "BLZ")
    _warm(client, "NZL", "BLZ")
    out = client.get("/database/records?nationality=NZL&destination=BLZ",
                     headers=ADMIN).json()
    assert out["summary"]["total"] == 2        # two products, one route
    rec = out["records"][0]
    assert rec["travel_document_country"] == "NZL"
    assert rec["field_status"]["visa_requirement"] == "filled"
    assert out["summary"]["source_coverage"] is not None
    # Filters combine: a requirement filter that matches nothing.
    none = client.get("/database/records?nationality=NZL&destination=BLZ"
                      "&requirement=Visa-free", headers=ADMIN).json()
    assert none["summary"]["total"] == 0
    # Readers cannot see the ops surface.
    assert client.get("/database/records", headers=READER).status_code == 403


def test_the_change_log_records_the_engine_answer(client):
    _warm(client, "NZL", "VUT")
    out = client.get("/database/changes?q=VUT", headers=ADMIN).json()
    assert any(c["action"] == "add" and c["origin"] == "engine"
               and (c["route"] or {}).get("destination_country") == "VUT"
               for c in out["changes"])
    assert client.get("/database/changes", headers=READER).status_code == 403


def test_excel_export_has_two_sheets_and_the_data(client):
    _warm(client, "NZL", "FSM")
    r = client.get("/database/export.xlsx?nationality=NZL&destination=FSM",
                   headers=ADMIN)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    # Data first: the acceptance standard reads the 25-field header row off
    # sheet 1; the descriptions ride second.
    assert wb.sheetnames == ["Data", "Field descriptions"]
    data = wb["Data"]
    header = [c.value for c in data[1]]
    assert header == list(tstation.FIELD_ORDER)
    assert data.max_row >= 3                    # header + two products
    fields = wb["Field descriptions"]
    # snapshot row (5.2) + header row + 25 field rows
    assert fields.max_row == 2 + len(tstation.FIELD_ORDER)
    assert fields["A1"].value == "Snapshot (UTC)"
    assert client.get("/database/export.xlsx", headers=READER).status_code == 403


def test_discretionary_validity_maps_to_the_stay_bound():
    """"Set by the consulate" is not a parser failure — it is the truth that
    no fixed validity exists. Their own display standard writes these as
    "Up to N days (determined at issuance)", so the record carries the stay
    length as the upper bound; with no stay known it stays honestly empty."""
    route = {"passport_nationality": "CHN", "destination_country": "FRA",
             "travel_purpose": "tourism"}
    g = {"disposition": "VISA_REQUIRED", "confidence": "high",
         "visa_products": [
             {"type": "Short-stay Schengen C", "entry": "single",
              "validity": "Up to trip duration / consulate discretion",
              "max_stay_days": 90,
              "fee": {"amount": 90, "currency": "EUR"}},
             {"type": "Mystery visa", "entry": "single",
              "validity": "as granted", "max_stay_days": None, "fee": None},
         ]}
    rows = tstation.records_for_route(route, g)
    assert (rows[0]["validity_duration"], rows[0]["validity_unit"]) == (90, "Day")
    assert rows[1]["validity_duration"] is None      # no bound, no guess


def test_parenthesized_and_on_arrival_validities_read_definitionally():
    """"Six (6) months" and "One (1) to three (3) months" state their own
    figures with parentheses between digit and unit; a range reads at the
    number written beside the unit word. A visa-on-arrival product with no
    validity text starts when granted, so its validity is the granted stay."""
    route = {"passport_nationality": "HKG", "destination_country": "GAB",
             "travel_purpose": "tourism"}
    g = {"disposition": "VISA_REQUIRED", "confidence": "high",
         "visa_products": [
             {"type": "e-Visa, short stay", "entry": "single",
              "validity": "One (1) to three (3) months", "max_stay_days": 90,
              "fee": {"amount": 70, "currency": "EUR"}},
             {"type": "e-Visa, long stay", "entry": "multiple",
              "validity": "Six (6) months", "max_stay_days": 90,
              "fee": {"amount": 185, "currency": "EUR"}},
             {"type": "Tourist visa on arrival (T)", "entry": "single",
              "validity": None, "max_stay_days": 30, "fee": None},
         ]}
    rows = tstation.records_for_route(route, g)
    assert (rows[0]["validity_duration"], rows[0]["validity_unit"]) == (3, "Month")
    assert (rows[1]["validity_duration"], rows[1]["validity_unit"]) == (6, "Month")
    assert (rows[2]["validity_duration"], rows[2]["validity_unit"]) == (30, "Day")


def test_spelled_out_validities_read_definitionally():
    """"Three months from issue" and "not exceeding five years" state their
    figures in words. Reading a written-out number is reading, not guessing."""
    route = {"passport_nationality": "USA", "destination_country": "GIN",
             "travel_purpose": "tourism"}
    g = {"disposition": "VISA_REQUIRED", "confidence": "high",
         "visa_products": [
             {"type": "Tourist / Entry e-Visa", "entry": "multiple",
              "validity": "not exceeding five years", "max_stay_days": 90,
              "fee": None},
             {"type": "Tourist visa", "entry": "single",
              "validity": "Three months from issue", "max_stay_days": 30,
              "fee": None},
         ]}
    rows = tstation.records_for_route(route, g)
    assert (rows[0]["validity_duration"], rows[0]["validity_unit"]) == (5, "Year")
    assert (rows[1]["validity_duration"], rows[1]["validity_unit"]) == (3, "Month")


def test_records_listing_serves_one_answer_per_route(client, db):
    """A route cached again per transit itinerary or arrival month changes
    the advice, not the product table. A stale variant with its own product
    names must not stand beside the fresh answer as phantom twins."""
    from datetime import datetime, timedelta, timezone
    from app.visa_snapshot.models import KimiRouteGuidanceCache
    for row in db.query(KimiRouteGuidanceCache).all():
        db.delete(row)
    db.commit()
    _warm(client, "NZL", "BLZ")
    fresh = db.query(KimiRouteGuidanceCache).one()
    stale_guidance = dict(fresh.guidance)
    stale_guidance["visa_products"] = [
        {"type": "Old-name visitor visa", "entry": "single",
         "validity": None, "max_stay_days": None, "fee": None}]
    db.add(KimiRouteGuidanceCache(
        cache_key=fresh.cache_key + "|via:JPN", route=dict(fresh.route),
        status="KIMI_PRIMARY", guidance=stale_guidance,
        generated_at=datetime.now(timezone.utc) - timedelta(days=3)))
    db.commit()
    out = client.get("/database/records?nationality=NZL&destination=BLZ",
                     headers=ADMIN).json()
    names = {r["visa_type_name"] for r in out["records"]}
    assert "Old-name visitor visa" not in names
    assert out["summary"]["total"] == 2


def test_proven_free_visas_keep_their_zero_and_get_a_currency():
    """"Exempt", "Nil" and "gratis" are how official pages say free. A zero
    backed by any of those words survives, and a surviving zero carries USD
    like the visa-exempt branch, so the currency cell never reads missing."""
    route = {"passport_nationality": "CHN", "destination_country": "NPL",
             "travel_purpose": "tourism"}
    g = {"disposition": "VISA_ON_ARRIVAL", "confidence": "high",
         "visa_products": [
             {"type": "30-day tourist visa", "entry": "multiple",
              "validity": "30 days", "max_stay_days": 30,
              "fee": {"amount": 0},
              "notes": "Gratis for Chinese nationals"},
             {"type": "Child visa", "entry": "single", "validity": "30 days",
              "max_stay_days": 30, "fee": {"amount": 0, "currency": "EUR"},
              "notes": "Children are exempt from the visa fee"},
             {"type": "Suspicious free visa", "entry": "single",
              "validity": "30 days", "max_stay_days": 30,
              "fee": {"amount": 0}, "notes": "No explanation given"},
         ]}
    rows = tstation.records_for_route(route, g)
    assert (rows[0]["visa_fee_amount"], rows[0]["visa_fee_currency"]) == (0, "USD")
    assert (rows[1]["visa_fee_amount"], rows[1]["visa_fee_currency"]) == (0, "EUR")
    assert rows[2]["visa_fee_amount"] is None     # an unexplained zero stays out


def test_productless_visa_routes_show_a_validity_bound():
    """A Senegalese tourist bound for France saw an empty validity cell: a
    product-less VISA_REQUIRED answer never set the column. The granted stay
    is the honest bound, same as product rows and the visa-free branch."""
    route = {"passport_nationality": "SEN", "destination_country": "FRA",
             "travel_purpose": "tourism"}
    g = {"disposition": "VISA_REQUIRED", "confidence": "high",
         "visa_category": "Short-stay Schengen C",
         "permitted_stay": "90 days in any 180-day period"}
    rows = tstation.records_for_route(route, g)
    assert (rows[0]["validity_duration"], rows[0]["validity_unit"]) == (90, "Day")
